#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
规则生成器 - 可视化对话框
通过框选参考行和目标单元格，智能生成内容修改规则
支持按行匹配（首件/成品）和按列匹配（过程）
"""

import re
import os
import sys
import uuid
from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLabel, QListWidget, QGroupBox, QCheckBox,
    QSplitter, QMessageBox, QStyledItemDelegate, QHeaderView, QComboBox
)
from PyQt5.QtCore import Qt, QRect
from PyQt5.QtGui import QColor, QBrush, QPainter, QPen

try:
    import win32com.client as win32
    WIN32_AVAILABLE = True
except ImportError:
    WIN32_AVAILABLE = False

import openpyxl

from core.log_manager import get_logger


# ============================================================
# 委托：绘制斜线边框
# ============================================================
class SlashDelegate(QStyledItemDelegate):
    """根据单元格的 UserRole 标记绘制对角线边框"""
    def paint(self, painter, option, index):
        super().paint(painter, option, index)
        if index.data(Qt.UserRole):
            painter.save()
            painter.setRenderHint(QPainter.Antialiasing)
            rect = option.rect
            pen = QPen(Qt.black, 1)
            painter.setPen(pen)
            painter.drawLine(rect.topLeft(), rect.bottomRight())
            painter.restore()


# ============================================================
# 规则生成器对话框
# ============================================================
class RuleGeneratorDialog(QDialog):
    def __init__(self, template_path: str, product_name: str = "", template_type: str = "", parent=None):
        super().__init__(parent)

        # ============================================================
        # 初始化日志
        # ============================================================
        self.logger = get_logger('RuleGenerator')
        self.logger.info("=" * 50)
        self.logger.info("规则生成器启动")
        self.logger.info(f"  模板路径: {template_path}")
        self.logger.info(f"  产品名称: {product_name}")
        self.logger.info(f"  模板类型: {template_type}")
        self.logger.info("=" * 50)

        # ============================================================
        # 窗口设置
        # ============================================================
        self.setWindowFlags(self.windowFlags() | Qt.WindowMaximizeButtonHint)
        self.setSizeGripEnabled(True)
        self.setWindowTitle("🧠 智能规则生成器")
        self.setGeometry(50, 50, 1200, 700)

        self.template_path = template_path
        self.product_name = product_name
        self.template_type = template_type  # 首件/过程/成品
        self.parent_window = parent

        # 数据
        self.rules = []
        self.reference_rows = {}  # {row_index: {'min': float, 'max': float, 'text': str, 'col': int}}
        self.reference_cols = {}  # {col_index: {'min': float, 'max': float, 'text': str, 'row': int}}  # 按列匹配用
        self.target_cells = []    # [(row, col), ...]
        self.match_mode = 'row'   # 'row' 或 'column'
        self.current_sheet_name = ""
        self.current_sheet_index = 0
        self.sheet_names = []
        self.wb = None
        self.excel_app = None
        self.worksheet = None
        self.all_worksheets = []

        self.logger.info("初始化 UI...")
        self.init_ui()
        self.logger.info("UI 初始化完成")

        # 根据模板类型自动设置匹配模式
        if template_type == '过程':
            self.match_mode = 'column'
            self.match_mode_combo.setCurrentText("按列匹配")
            self.logger.info("自动切换到按列匹配模式（过程模板）")
        else:
            self.match_mode = 'row'
            self.match_mode_combo.setCurrentText("按行匹配")
            self.logger.info("自动切换到按行匹配模式（首件/成品模板）")

        self.logger.info("开始加载 Excel...")
        self.load_excel()

    # ============================================================
    # 界面创建
    # ============================================================
    def init_ui(self):
        self.logger.debug("init_ui: 开始")

        layout = QVBoxLayout(self)

        # ========== 工具栏 ==========
        toolbar = QHBoxLayout()

        self.lbl_file = QLabel(f"📂 {os.path.basename(self.template_path)}")
        toolbar.addWidget(self.lbl_file)

        toolbar.addWidget(QLabel("  工作表:"))
        self.sheet_combo = QComboBox()
        self.sheet_combo.setMinimumWidth(180)
        self.sheet_combo.currentIndexChanged.connect(self._on_sheet_changed)
        toolbar.addWidget(self.sheet_combo)

        toolbar.addWidget(QLabel("  匹配模式:"))
        self.match_mode_combo = QComboBox()
        self.match_mode_combo.addItems(["按行匹配", "按列匹配"])
        self.match_mode_combo.setMinimumWidth(120)
        self.match_mode_combo.currentTextChanged.connect(self._on_match_mode_changed)
        toolbar.addWidget(self.match_mode_combo)

        toolbar.addStretch()

        self.btn_ref_mode = QPushButton("📐 选择参考")
        self.btn_ref_mode.setCheckable(True)
        self.btn_ref_mode.clicked.connect(lambda: self.set_mode('reference'))
        toolbar.addWidget(self.btn_ref_mode)

        self.btn_target_mode = QPushButton("🎯 选择目标")
        self.btn_target_mode.setCheckable(True)
        self.btn_target_mode.clicked.connect(lambda: self.set_mode('target'))
        toolbar.addWidget(self.btn_target_mode)

        self.btn_generate = QPushButton("➕ 生成规则")
        self.btn_generate.clicked.connect(self.generate_rules)
        self.btn_generate.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")
        toolbar.addWidget(self.btn_generate)

        self.btn_import = QPushButton("📥 导入规则")
        self.btn_import.clicked.connect(self.import_rules)
        self.btn_import.setStyleSheet("background-color: #2196F3; color: white; font-weight: bold;")
        toolbar.addWidget(self.btn_import)

        self.btn_close = QPushButton("✕ 关闭")
        self.btn_close.clicked.connect(self.reject)
        toolbar.addWidget(self.btn_close)

        layout.addLayout(toolbar)

        # ========== 状态栏 ==========
        status_layout = QHBoxLayout()
        self.status_label = QLabel("💡 提示：先框选参考，再框选目标，然后点击生成规则")
        self.status_label.setStyleSheet("background-color: #f0f0f0; padding: 5px; border: 1px solid #ccc;")
        status_layout.addWidget(self.status_label, 1)

        self.chk_merge = QCheckBox("合并同列目标为范围 (按列匹配时)")
        self.chk_merge.setChecked(True)
        status_layout.addWidget(self.chk_merge)

        layout.addLayout(status_layout)

        # ========== 表格 ==========
        self.table = QTableWidget()
        self.table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.table.setSelectionBehavior(QTableWidget.SelectItems)
        self.table.itemSelectionChanged.connect(self.on_selection_changed)
        self.table.itemDoubleClicked.connect(self.on_cell_double_click)
        self.table.setItemDelegate(SlashDelegate())
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        layout.addWidget(self.table, 5)

        # ========== 底部：参考列表 + 规则列表 ==========
        bottom_splitter = QSplitter(Qt.Horizontal)

        ref_group = QGroupBox("📋 已选参考")
        ref_layout = QVBoxLayout()
        self.ref_list_widget = QListWidget()
        ref_layout.addWidget(self.ref_list_widget)
        ref_group.setLayout(ref_layout)
        bottom_splitter.addWidget(ref_group)

        rule_group = QGroupBox("📄 已生成规则 (双击删除)")
        rule_layout = QVBoxLayout()
        self.rule_list = QListWidget()
        self.rule_list.itemDoubleClicked.connect(self.delete_rule)
        rule_layout.addWidget(self.rule_list)
        rule_group.setLayout(rule_layout)
        bottom_splitter.addWidget(rule_group)

        bottom_splitter.setSizes([300, 600])
        layout.addWidget(bottom_splitter, 2)

        self.set_mode(None)
        self.logger.debug("init_ui: 完成")

    # ============================================================
    # 模式切换
    # ============================================================
    def set_mode(self, mode):
        self.mode = mode
        self.btn_ref_mode.setChecked(mode == 'reference')
        self.btn_target_mode.setChecked(mode == 'target')

        if mode == 'reference':
            if self.match_mode == 'row':
                self.status_label.setText("🔍 请框选【包含公差描述】的连续行（如 C8:C15），每行将提取公差")
            else:
                self.status_label.setText("🔍 请框选【包含公差描述】的连续列（如 D5:J5），每列将提取公差")
            self.table.setSelectionMode(QTableWidget.ContiguousSelection)
            self.table.clearSelection()
        elif mode == 'target':
            self.status_label.setText("🎯 请框选【要填充随机数】的单元格（可跨列、跨行），将按当前匹配模式匹配参考")
            self.table.setSelectionMode(QTableWidget.ExtendedSelection)
            self.table.clearSelection()
        else:
            self.status_label.setText("💡 已退出选择模式")
            self.table.setSelectionMode(QTableWidget.ExtendedSelection)

    # ============================================================
    # 匹配模式切换
    # ============================================================
    def _on_match_mode_changed(self, text):
        if text == "按行匹配":
            self.match_mode = 'row'
            self.logger.info("切换到按行匹配模式")
        else:
            self.match_mode = 'column'
            self.logger.info("切换到按列匹配模式")

        # 清空已选参考
        self.reference_rows.clear()
        self.reference_cols.clear()
        self.ref_list_widget.clear()
        self.target_cells.clear()
        self.status_label.setText(f"💡 已切换到 {text}，请重新选择参考")

    # ============================================================
    # 获取合并区域（兼容 Excel 和 WPS）
    # ============================================================
    def _get_merge_areas(self, sheet):
        merge_areas = []

        try:
            if hasattr(sheet.Cells, 'MergeAreas'):
                areas = sheet.Cells.MergeAreas
                if areas:
                    for area in areas:
                        merge_areas.append({
                            'Row': area.Row,
                            'Column': area.Column,
                            'Rows': area.Rows,
                            'Columns': area.Columns
                        })
                    self.logger.info(f"通过 MergeAreas 获取到 {len(merge_areas)} 个合并区域")
                    return merge_areas
        except Exception as e:
            self.logger.debug(f"MergeAreas 方式失败: {e}")

        try:
            used_range = sheet.UsedRange
            rows = used_range.Rows.Count
            cols = used_range.Columns.Count

            if rows > 200:
                rows = 200
            if cols > 30:
                cols = 30

            self.logger.debug(f"遍历 UsedRange ({rows}x{cols}) 查找合并区域...")
            processed = set()

            for r in range(1, rows + 1):
                for c in range(1, cols + 1):
                    try:
                        cell = sheet.Cells(r, c)
                        if hasattr(cell, 'MergeCells'):
                            is_merged = cell.MergeCells
                            if is_merged:
                                if hasattr(cell, 'MergeArea'):
                                    area = cell.MergeArea
                                    key = (area.Row, area.Column)
                                    if key not in processed:
                                        processed.add(key)
                                        merge_areas.append({
                                            'Row': area.Row,
                                            'Column': area.Column,
                                            'Rows': area.Rows,
                                            'Columns': area.Columns
                                        })
                                        self.logger.debug(f"  找到合并区域: 行{area.Row} 列{area.Column} 跨{area.Rows.Count}行 {area.Columns.Count}列")
                        elif hasattr(cell, 'MergeArea'):
                            try:
                                area = cell.MergeArea
                                key = (area.Row, area.Column)
                                if key not in processed:
                                    processed.add(key)
                                    merge_areas.append({
                                        'Row': area.Row,
                                        'Column': area.Column,
                                        'Rows': area.Rows,
                                        'Columns': area.Columns
                                    })
                            except:
                                pass
                    except Exception:
                        pass

            if merge_areas:
                self.logger.info(f"通过遍历方式获取到 {len(merge_areas)} 个合并区域")
            else:
                self.logger.debug("未找到任何合并区域")

            return merge_areas

        except Exception as e:
            self.logger.warning(f"遍历查找合并区域失败: {e}")

        return merge_areas

    # ============================================================
    # 加载 Excel
    # ============================================================
    def load_excel(self):
        self.logger.info("load_excel: 开始")

        if not self.template_path or not os.path.exists(self.template_path):
            self.logger.error(f"模板文件不存在: {self.template_path}")
            self.status_label.setText("❌ 模板文件不存在")
            return

        self.logger.info(f"使用 win32com 加载: {self.template_path}")

        if WIN32_AVAILABLE:
            try:
                self.logger.info("启动 Excel 应用程序...")
                excel_app = win32.Dispatch("Excel.Application")
                excel_app.Visible = False
                excel_app.DisplayAlerts = False
                self.logger.info("Excel 已启动")

                self.logger.info(f"打开工作簿: {self.template_path}")
                wb = excel_app.Workbooks.Open(self.template_path)
                self.logger.info("工作簿已打开")

                self.wb = wb
                self.excel_app = excel_app

                self.all_worksheets = []
                self.sheet_names = []
                for sheet in wb.Worksheets:
                    self.all_worksheets.append(sheet)
                    self.sheet_names.append(sheet.Name)
                self.logger.info(f"共 {len(self.sheet_names)} 个工作表: {self.sheet_names}")

                self.current_sheet_index = 0
                self.current_sheet_name = self.sheet_names[0]
                self.worksheet = self.all_worksheets[0]

                self.sheet_combo.blockSignals(True)
                self.sheet_combo.clear()
                self.sheet_combo.addItems(self.sheet_names)
                self.sheet_combo.setCurrentIndex(0)
                self.sheet_combo.blockSignals(False)

                self._load_sheet_data(self.worksheet)

                self.status_label.setText(f"✅ 加载成功，当前工作表: {self.current_sheet_name}")
                self.table.resizeColumnsToContents()
                self.table.resizeRowsToContents()
                self.logger.info("load_excel: 完成")
                return

            except Exception as e:
                self.logger.error(f"win32com 加载失败: {e}")
                import traceback
                self.logger.error(traceback.format_exc())

                try:
                    if self.wb:
                        self.wb.Close(SaveChanges=False)
                        self.logger.debug("工作簿已关闭")
                except:
                    pass
                try:
                    if self.excel_app:
                        self.excel_app.Quit()
                        self.logger.debug("Excel 已退出")
                except:
                    pass
                self.wb = None
                self.excel_app = None
                self.worksheet = None
                self.status_label.setText(f"⚠️ win32com失败，改用 openpyxl")

        self.logger.info("使用 openpyxl 加载...")
        try:
            self.load_with_openpyxl(self.template_path)
            self.logger.info("openpyxl 加载完成")
        except Exception as e:
            self.logger.error(f"openpyxl 加载失败: {e}")
            import traceback
            self.logger.error(traceback.format_exc())
            QMessageBox.critical(self, "加载错误", f"所有加载方式均失败：{str(e)}")

    # ============================================================
    # 加载指定工作表的数据
    # ============================================================
    def _load_sheet_data(self, sheet):
        self.logger.info(f"_load_sheet_data: 加载工作表 {sheet.Name}")

        self.reference_rows.clear()
        self.reference_cols.clear()
        self.target_cells.clear()
        self.rules.clear()
        self.rule_list.clear()
        self.ref_list_widget.clear()
        self.table.clearSelection()

        self.current_sheet_name = sheet.Name
        self.worksheet = sheet

        merge_areas = self._get_merge_areas(sheet)
        self.logger.info(f"检测到 {len(merge_areas) if merge_areas else 0} 个合并区域")

        max_row, max_col = self._scan_data_range(sheet, merge_areas)
        self.logger.info(f"实际数据范围: {max_row} 行, {max_col} 列")

        self._populate_table_from_win32(sheet, max_row, max_col, merge_areas)

        self.status_label.setText(f"✅ 当前工作表: {self.current_sheet_name}，共 {max_row} 行，{max_col} 列")
        self.table.resizeColumnsToContents()
        self.table.resizeRowsToContents()

    # ============================================================
    # 扫描数据范围
    # ============================================================
    def _scan_data_range(self, sheet, merge_areas):
        merge_rows = set()
        merge_cols = set()
        if merge_areas:
            for area in merge_areas:
                for r in range(area['Row'], area['Row'] + area['Rows'].Count):
                    merge_rows.add(r)
                for c in range(area['Column'], area['Column'] + area['Columns'].Count):
                    merge_cols.add(c)

        max_row = 0
        max_col = 0

        for r in range(1, 201):
            row_has_data = False
            for c in range(1, 31):
                val = sheet.Cells(r, c).Text
                if val and val.strip():
                    row_has_data = True
                    if c > max_col:
                        max_col = c
                    break
            if not row_has_data and r in merge_rows:
                row_has_data = True
                for c in merge_cols:
                    if c > max_col:
                        max_col = c

            if row_has_data:
                max_row = r
            else:
                empty_count = 0
                for check_r in range(r, min(r + 5, 201)):
                    has_data = False
                    for c in range(1, 31):
                        val = sheet.Cells(check_r, c).Text
                        if val and val.strip():
                            has_data = True
                            break
                    if not has_data and check_r not in merge_rows:
                        empty_count += 1
                    else:
                        empty_count = 0
                if empty_count >= 5:
                    break

        if max_row == 0:
            used_range = sheet.UsedRange
            max_row = used_range.Rows.Count
            max_col = used_range.Columns.Count
            if max_col > 30:
                max_col = 30
        else:
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 5):
                    if c > 30:
                        break
                    val = sheet.Cells(r, c).Text
                    if val and val.strip() and c > max_col:
                        max_col = c
            if merge_cols:
                max_merge_col = max(merge_cols)
                if max_merge_col > max_col:
                    max_col = min(max_merge_col, 30)
            if max_col > 30:
                max_col = 30

        if max_col == 0:
            max_col = 16

        return max_row, max_col

    # ============================================================
    # 填充表格
    # ============================================================
    def _populate_table_from_win32(self, sheet, max_row, max_col, merge_areas):
        self.logger.info(f"_populate_table_from_win32: 开始填充 {max_row}x{max_col}")

        self.table.setRowCount(max_row)
        self.table.setColumnCount(max_col)

        for c in range(max_col):
            self.table.setHorizontalHeaderItem(c, QTableWidgetItem(chr(65 + c)))

        xlDiagonalDown = 5
        xlDiagonalUp = 6
        xlLineStyleNone = -4142

        self.logger.info("读取单元格数据...")
        for r in range(max_row):
            if r % 10 == 0:
                self.logger.debug(f"  读取第 {r+1}/{max_row} 行")
            for c in range(max_col):
                cell = sheet.Cells(r + 1, c + 1)
                val = cell.Text if cell.Text else None
                display = str(val) if val is not None else ""
                item = QTableWidgetItem(display)
                item.setFlags(item.flags() | Qt.ItemIsSelectable | Qt.ItemIsEnabled)

                try:
                    font = cell.Font
                    if font.Bold:
                        f = item.font()
                        f.setBold(True)
                        item.setFont(f)
                    if font.Italic:
                        f = item.font()
                        f.setItalic(True)
                        item.setFont(f)
                except:
                    pass

                if val and isinstance(val, str):
                    if "（" in val and "-" in val and "）" in val:
                        item.setBackground(QBrush(QColor(255, 255, 200)))

                has_diagonal = False
                try:
                    border_down = cell.Borders(xlDiagonalDown)
                    border_up = cell.Borders(xlDiagonalUp)
                    if border_down.LineStyle != xlLineStyleNone:
                        has_diagonal = True
                    if border_up.LineStyle != xlLineStyleNone:
                        has_diagonal = True
                except Exception:
                    pass

                if has_diagonal:
                    item.setData(Qt.UserRole, True)

                self.table.setItem(r, c, item)

        self.logger.info("单元格数据读取完成")

        if merge_areas:
            count = 0
            for area in merge_areas:
                first_row = area['Row'] - 1
                first_col = area['Column'] - 1
                row_span = area['Rows'].Count
                col_span = area['Columns'].Count

                if first_row < max_row and first_col < max_col:
                    actual_row_span = min(row_span, max_row - first_row)
                    actual_col_span = min(col_span, max_col - first_col)
                    if actual_row_span > 1 or actual_col_span > 1:
                        self.table.setSpan(first_row, first_col, actual_row_span, actual_col_span)
                        count += 1
            self.logger.info(f"已应用 {count} 个合并区域")
        else:
            self.logger.info("没有合并区域需要处理")

        self.table.setShowGrid(True)
        self.table.setGridStyle(Qt.SolidLine)
        self.logger.info("_populate_table_from_win32: 完成")

    # ============================================================
    # 工作表切换
    # ============================================================
    def _on_sheet_changed(self, index):
        if index < 0 or index >= len(self.all_worksheets):
            return

        if self.reference_rows or self.reference_cols or self.target_cells or self.rules:
            reply = QMessageBox.question(
                self,
                "切换工作表",
                "切换工作表将清空当前已选的参考、目标单元格和已生成的规则。\n\n确定要切换吗？",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply == QMessageBox.No:
                self.sheet_combo.blockSignals(True)
                self.sheet_combo.setCurrentIndex(self.current_sheet_index)
                self.sheet_combo.blockSignals(False)
                return

        self.current_sheet_index = index
        self.worksheet = self.all_worksheets[index]
        self._load_sheet_data(self.worksheet)
        self.set_mode(None)
        self.logger.info(f"切换到工作表: {self.current_sheet_name}")

    # ============================================================
    # openpyxl 加载（备用）
    # ============================================================
    def load_with_openpyxl(self, file_path):
        self.logger.info("load_with_openpyxl: 开始")
        wb = openpyxl.load_workbook(file_path, data_only=True)
        self.wb_openpyxl = wb

        self.sheet_names = wb.sheetnames
        self.logger.info(f"共 {len(self.sheet_names)} 个工作表: {self.sheet_names}")

        self.current_sheet_index = 0
        self.current_sheet_name = self.sheet_names[0]

        self.sheet_combo.blockSignals(True)
        self.sheet_combo.clear()
        self.sheet_combo.addItems(self.sheet_names)
        self.sheet_combo.setCurrentIndex(0)
        self.sheet_combo.blockSignals(False)

        self._load_openpyxl_sheet(self.sheet_names[0])
        self.logger.info("load_with_openpyxl: 完成")

    def _load_openpyxl_sheet(self, sheet_name):
        wb = self.wb_openpyxl
        sheet = wb[sheet_name]
        self.current_sheet_name = sheet_name

        merge_ranges = list(sheet.merged_cells.ranges)
        merge_rows = set()
        merge_cols = set()
        for mr in merge_ranges:
            for r in range(mr.min_row, mr.max_row + 1):
                merge_rows.add(r)
            for c in range(mr.min_col, mr.max_col + 1):
                merge_cols.add(c)

        max_row = 0
        max_col = 0
        for r in range(1, min(sheet.max_row + 1, 201)):
            has_data = False
            for c in range(1, min(sheet.max_column + 1, 31)):
                val = sheet.cell(row=r, column=c).value
                if val and str(val).strip():
                    has_data = True
                    if c > max_col:
                        max_col = c
                    break
            if not has_data and r in merge_rows:
                has_data = True
                for c in merge_cols:
                    if c > max_col:
                        max_col = c
            if has_data:
                max_row = r
            else:
                empty_count = 0
                for check_r in range(r, min(r + 5, 201)):
                    has_data2 = False
                    for c in range(1, min(sheet.max_column + 1, 31)):
                        val2 = sheet.cell(row=check_r, column=c).value
                        if val2 and str(val2).strip():
                            has_data2 = True
                            break
                    if not has_data2 and check_r not in merge_rows:
                        empty_count += 1
                    else:
                        empty_count = 0
                if empty_count >= 5:
                    break

        if max_col == 0:
            max_col = min(sheet.max_column, 30)
        if max_col > 30:
            max_col = 30

        self.logger.info(f"openpyxl 实际数据: {max_row} 行, {max_col} 列")

        self.table.setRowCount(max_row)
        self.table.setColumnCount(max_col)
        for c in range(max_col):
            self.table.setHorizontalHeaderItem(c, QTableWidgetItem(chr(65 + c)))

        for r in range(max_row):
            if r % 10 == 0:
                self.logger.debug(f"  读取第 {r+1}/{max_row} 行")
            for c in range(max_col):
                cell = sheet.cell(row=r+1, column=c+1)
                val = cell.value
                display = str(val) if val is not None else ""
                item = QTableWidgetItem(display)
                item.setFlags(item.flags() | Qt.ItemIsSelectable | Qt.ItemIsEnabled)

                if cell.has_style:
                    font = cell.font
                    if font:
                        if font.bold:
                            f = item.font()
                            f.setBold(True)
                            item.setFont(f)
                        if font.italic:
                            f = item.font()
                            f.setItalic(True)
                            item.setFont(f)

                if val and isinstance(val, str) and "（" in val and "-" in val and "）" in val:
                    item.setBackground(QBrush(QColor(255, 255, 200)))

                border = cell.border
                if border and (border.diagonal or border.diagonalUp or border.diagonalDown):
                    item.setData(Qt.UserRole, True)

                self.table.setItem(r, c, item)

        for mr in merge_ranges:
            first_row = mr.min_row - 1
            first_col = mr.min_col - 1
            row_span = mr.max_row - mr.min_row + 1
            col_span = mr.max_col - mr.min_col + 1
            if first_row < max_row and first_col < max_col:
                actual_row_span = min(row_span, max_row - first_row)
                actual_col_span = min(col_span, max_col - first_col)
                if actual_row_span > 1 or actual_col_span > 1:
                    self.table.setSpan(first_row, first_col, actual_row_span, actual_col_span)

        self.table.setShowGrid(True)
        self.table.setGridStyle(Qt.SolidLine)
        self.table.resizeColumnsToContents()
        self.table.resizeRowsToContents()

        self.reference_rows.clear()
        self.reference_cols.clear()
        self.target_cells.clear()
        self.rules.clear()
        self.rule_list.clear()
        self.ref_list_widget.clear()
        self.table.clearSelection()

    # ============================================================
    # 交互事件
    # ============================================================
    def on_cell_double_click(self, item):
        if not item:
            return
        text = item.text()
        if "（" in text and "-" in text and "）" in text:
            row, col = item.row(), item.column()
            self.process_reference(row, col, text)
        else:
            QMessageBox.information(self, "提示", "当前单元格不包含公差信息")

    def on_selection_changed(self):
        selected = self.table.selectedIndexes()
        if not selected:
            return

        if self.mode == 'reference':
            if self.match_mode == 'row':
                rows = set(idx.row() for idx in selected)
                for row in rows:
                    for idx in selected:
                        if idx.row() == row:
                            item = self.table.item(idx.row(), idx.column())
                            if item:
                                text = item.text()
                                if "（" in text and "-" in text and "）" in text:
                                    self.process_reference(row, idx.column(), text)
                                    break
                    else:
                        self.status_label.setText(f"⚠️ 第{row+1}行未找到公差描述")
            else:
                # 按列匹配：提取每列的公差
                cols = set(idx.column() for idx in selected)
                for col in cols:
                    for idx in selected:
                        if idx.column() == col:
                            item = self.table.item(idx.row(), idx.column())
                            if item:
                                text = item.text()
                                if "（" in text and "-" in text and "）" in text:
                                    self.process_reference_col(idx.row(), col, text)
                                    break
                    else:
                        self.status_label.setText(f"⚠️ 第{chr(65+col)}列未找到公差描述")

        elif self.mode == 'target':
            self.target_cells = [(idx.row(), idx.column()) for idx in selected]
            self.status_label.setText(f"🎯 已选中 {len(self.target_cells)} 个目标单元格")

    def process_reference(self, row, col, text):
        match = re.search(r"[（(]([\d.]+)-([\d.]+)[）)]", text)
        if match:
            min_val = float(match.group(1))
            max_val = float(match.group(2))
            self.reference_rows[row] = {'min': min_val, 'max': max_val, 'text': text, 'col': col}
            self.ref_list_widget.clear()
            for r, info in self.reference_rows.items():
                cell_pos = f"{chr(65 + info['col'])}{r+1}"
                self.ref_list_widget.addItem(
                    f"行{r+1}  {cell_pos}  {info['min']}~{info['max']}  {info['text'][:30]}"
                )
            self.status_label.setText(f"✅ 已添加参考行 {row+1} : {min_val}~{max_val}")
        else:
            QMessageBox.warning(self, "提取失败", "未找到 (数字-数字) 格式的公差范围")

    def process_reference_col(self, row, col, text):
        match = re.search(r"[（(]([\d.]+)-([\d.]+)[）)]", text)
        if match:
            min_val = float(match.group(1))
            max_val = float(match.group(2))
            self.reference_cols[col] = {'min': min_val, 'max': max_val, 'text': text, 'row': row}
            self.ref_list_widget.clear()
            for c, info in self.reference_cols.items():
                cell_pos = f"{chr(65 + c)}{info['row']}"
                self.ref_list_widget.addItem(
                    f"列{chr(65+c)}  {cell_pos}  {info['min']}~{info['max']}  {info['text'][:30]}"
                )
            self.status_label.setText(f"✅ 已添加参考列 {chr(65+col)} : {min_val}~{max_val}")
        else:
            QMessageBox.warning(self, "提取失败", "未找到 (数字-数字) 格式的公差范围")

    # ============================================================
    # 生成规则
    # ============================================================
    def generate_rules(self):
        self.logger.info("generate_rules: 开始")

        if self.match_mode == 'row':
            if not self.reference_rows:
                QMessageBox.warning(self, "缺少参考", "请先选择至少一个参考行（公差行）")
                return
            if not self.target_cells:
                QMessageBox.warning(self, "缺少目标", "请先选择目标单元格")
                return

            self.rules = []
            self.rule_list.clear()
            self.logger.info(f"参考行: {len(self.reference_rows)} 个, 目标单元格: {len(self.target_cells)} 个")

            for ref_row, ref_info in self.reference_rows.items():
                same_row_targets = []
                for (r, c) in self.target_cells:
                    if r == ref_row:
                        same_row_targets.append(f"{chr(65 + c)}{r+1}")

                if not same_row_targets:
                    continue

                if len(same_row_targets) == 1:
                    target_str = same_row_targets[0]
                    target_type = 'cell'
                else:
                    cols = [ord(t[0]) - 65 for t in same_row_targets]
                    cols.sort()
                    is_continuous = all(cols[i+1] - cols[i] == 1 for i in range(len(cols) - 1))
                    if is_continuous and len(cols) > 1:
                        start_col = chr(65 + min(cols))
                        end_col = chr(65 + max(cols))
                        target_str = f"{start_col}{ref_row+1}-{end_col}{ref_row+1}"
                        target_type = 'range'
                    else:
                        target_str = ",".join(same_row_targets)
                        target_type = 'cells'

                rule = {
                    "id": str(uuid.uuid4())[:8],
                    "target_type": target_type,
                    "target": target_str,
                    "value_type": "random",
                    "sheet_name": self.current_sheet_name,
                    "min_val": ref_info['min'],
                    "max_val": ref_info['max'],
                    "decimals": 2,
                    "enabled": True
                }
                self.rules.append(rule)
                self.rule_list.addItem(
                    f"行{ref_row+1} → {target_str}  {ref_info['min']}~{ref_info['max']}"
                )
                self.logger.debug(f"生成规则: {target_str} ({ref_info['min']}~{ref_info['max']})")
            self.status_label.setText(f"✅ 生成 {len(self.rules)} 条规则")

        else:
            # ============================================================
            # 按列匹配（过程模板）
            # ============================================================
            if not self.reference_cols:
                QMessageBox.warning(self, "缺少参考", "请先选择至少一个参考列（公差列）")
                return
            if not self.target_cells:
                QMessageBox.warning(self, "缺少目标", "请先选择目标单元格")
                return

            self.rules = []
            self.rule_list.clear()
            self.logger.info(f"参考列: {len(self.reference_cols)} 个, 目标单元格: {len(self.target_cells)} 个")

            # 按列分组目标单元格
            target_by_col = {}
            for (r, c) in self.target_cells:
                if c not in target_by_col:
                    target_by_col[c] = []
                target_by_col[c].append(r)

            for ref_col, ref_info in self.reference_cols.items():
                if ref_col not in target_by_col:
                    self.logger.debug(f"列 {chr(65+ref_col)} 没有目标单元格，跳过")
                    continue

                rows = target_by_col[ref_col]
                rows.sort()

                if len(rows) == 1:
                    target_str = f"{chr(65 + ref_col)}{rows[0]+1}"
                    target_type = 'cell'
                else:
                    # 检查行是否连续
                    is_continuous = all(rows[i+1] - rows[i] == 1 for i in range(len(rows) - 1))
                    if is_continuous and len(rows) > 1:
                        target_str = f"{chr(65 + ref_col)}{rows[0]+1}-{chr(65 + ref_col)}{rows[-1]+1}"
                        target_type = 'column'
                    else:
                        # 离散行
                        addresses = [f"{chr(65 + ref_col)}{r+1}" for r in rows]
                        target_str = ",".join(addresses)
                        target_type = 'cells'

                rule = {
                    "id": str(uuid.uuid4())[:8],
                    "target_type": target_type,
                    "target": target_str,
                    "value_type": "random",
                    "sheet_name": self.current_sheet_name,
                    "min_val": ref_info['min'],
                    "max_val": ref_info['max'],
                    "decimals": 2,
                    "enabled": True
                }
                self.rules.append(rule)
                self.rule_list.addItem(
                    f"列{chr(65+ref_col)} → {target_str}  {ref_info['min']}~{ref_info['max']}"
                )
                self.logger.debug(f"生成规则: {target_str} ({ref_info['min']}~{ref_info['max']})")

            self.status_label.setText(f"✅ 生成 {len(self.rules)} 条规则")

        self.logger.info(f"generate_rules: 完成，共 {len(self.rules)} 条规则")
        self.table.clearSelection()
        self.set_mode(None)

    def delete_rule(self, item):
        row = self.rule_list.row(item)
        self.rule_list.takeItem(row)
        del self.rules[row]
        self.status_label.setText("🗑️ 已删除规则")

    # ============================================================
    # 导入规则到主程序
    # ============================================================
    def import_rules(self):
        self.logger.info("import_rules: 开始")
        if not self.rules:
            QMessageBox.warning(self, "错误", "规则列表为空，请先生成规则")
            return

        if not self.parent_window:
            QMessageBox.warning(self, "错误", "无法连接到主程序")
            return

        existing_targets = [r.target for r in self.parent_window.rules] if hasattr(self.parent_window, 'rules') else []
        self.logger.info(f"主程序现有规则: {len(existing_targets)} 条")

        new_rules_count = 0
        skipped_count = 0

        for rule in self.rules:
            if rule['target'] in existing_targets:
                self.logger.debug(f"跳过重复规则: {rule['target']}")
                skipped_count += 1
                continue

            from models.config_models import Rule
            new_rule = Rule(
                id=rule['id'],
                target_type=rule['target_type'],
                target=rule['target'],
                value_type=rule['value_type'],
                sheet_name=rule['sheet_name'],
                min_val=rule['min_val'],
                max_val=rule['max_val'],
                decimals=rule['decimals'],
                enabled=rule['enabled']
            )
            self.parent_window.rules.append(new_rule)
            existing_targets.append(rule['target'])
            new_rules_count += 1
            self.logger.debug(f"导入规则: {rule['target']}")

        if new_rules_count > 0:
            self.parent_window._update_rule_list()
            self.parent_window._save_rules()

        msg = f"✅ 成功导入 {new_rules_count} 条规则"
        if skipped_count > 0:
            msg += f"，跳过 {skipped_count} 条重复规则"
        QMessageBox.information(self, "导入成功", msg)
        self.status_label.setText(msg)
        self.logger.info(f"import_rules: 完成，导入 {new_rules_count} 条，跳过 {skipped_count} 条")

        self.accept()

    # ============================================================
    # 关闭事件
    # ============================================================
    def closeEvent(self, event):
        self.logger.info("规则生成器关闭")
        try:
            if self.wb:
                self.wb.Close(SaveChanges=False)
                self.logger.debug("工作簿已关闭")
        except Exception as e:
            self.logger.warning(f"关闭工作簿失败: {e}")
        try:
            if self.excel_app:
                self.excel_app.Quit()
                self.logger.debug("Excel 已退出")
        except Exception as e:
            self.logger.warning(f"退出 Excel 失败: {e}")
        event.accept()